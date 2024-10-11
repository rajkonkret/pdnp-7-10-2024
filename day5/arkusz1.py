import openpyxl

# pip install openpyxl

# działania na plikach excel

workbook = openpyxl.load_workbook('sales.xlsx')
worksheet = workbook.active
print(worksheet.title)  # Arkusz1

lista = []
for i in worksheet:
    print(i)
# (<Cell 'Arkusz1'.A1>, <Cell 'Arkusz1'.B1>, <Cell 'Arkusz1'.C1>)
# (<Cell 'Arkusz1'.A2>, <Cell 'Arkusz1'.B2>, <Cell 'Arkusz1'.C2>)
# (<Cell 'Arkusz1'.A3>, <Cell 'Arkusz1'.B3>, <Cell 'Arkusz1'.C3>)
# (<Cell 'Arkusz1'.A4>, <Cell 'Arkusz1'.B4>, <Cell 'Arkusz1'.C4>)
# (<Cell 'Arkusz1'.A5>, <Cell 'Arkusz1'.B5>, <Cell 'Arkusz1'.C5>)

for i in range(worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        lista.append(col[i].value)

print(lista)
# ['Sales Date', 'Sales Person', 'Amount',
#  datetime.datetime(2018, 5, 12, 0, 0), 'Sila Ahmed', 60000, datetime.datetime(2019, 12, 6, 0, 0),
#  'Mir Hossain', 50000, datetime.datetime(2020, 8, 9, 0, 0), 'Sarmin Jahan', 45000, datetime.datetime(2021, 4, 7, 0, 0),
#  'Mahmudul Hasan', 30000]
for i in lista:
    print(i)
# Sales Person
# Amount
# 2018-05-12 00:00:00
# Sila Ahmed
# 60000
# 2019-12-06 00:00:00
# Mir Hossain
# 50000
# 2020-08-09 00:00:00
# Sarmin Jahan
# 45000
# 2021-04-07 00:00:00
# Mahmudul Hasan
# 30000
for i in range(0, len(lista), 3):
    print(lista[i: i + 3])
# ['Sales Date', 'Sales Person', 'Amount']
# [datetime.datetime(2018, 5, 12, 0, 0), 'Sila Ahmed', 60000]
# [datetime.datetime(2019, 12, 6, 0, 0), 'Mir Hossain', 50000]
# [datetime.datetime(2020, 8, 9, 0, 0), 'Sarmin Jahan', 45000]
# [datetime.datetime(2021, 4, 7, 0, 0), 'Mahmudul Hasan', 30000]